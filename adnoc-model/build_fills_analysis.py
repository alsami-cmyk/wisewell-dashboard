from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors & Styles ──
ADNOC_BLUE = "0047BA"
WHITE = "FFFFFF"
LIGHT_BLUE = "E8F0FB"
LIGHT_GRAY = "F5F5F5"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=ADNOC_BLUE)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

section_font = Font(name="Arial", bold=True, color=ADNOC_BLUE, size=12)
label_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
blue_input = Font(name="Arial", color="0000FF", size=10)  # Hardcoded inputs
black_formula = Font(name="Arial", color="000000", size=10)  # Formulas
result_font = Font(name="Arial", bold=True, color=ADNOC_BLUE, size=11)

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
bottom_border = Border(bottom=Side(style="medium", color=ADNOC_BLUE))
highlight_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
alt_fill = PatternFill("solid", fgColor=LIGHT_GRAY)

NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
DEC_FMT = '#,##0.0'

def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, num_fmt=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if num_fmt: cell.number_format = num_fmt

def set_cell(ws, row, col, value, font=None, fill=None, num_fmt=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if num_fmt: cell.number_format = num_fmt
    if alignment: cell.alignment = alignment
    return cell

# ════════════════════════════════════════
# SHEET 1: Fills Buildup
# ════════════════════════════════════════
ws1 = wb.active
ws1.title = "Fills Buildup"
ws1.sheet_properties.tabColor = ADNOC_BLUE

ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 18

# Title bar
ws1.merge_cells('A1:F1')
title_cell = set_cell(ws1, 1, 1, "ADNOC Oasis Water Refill — Fill Volume Analysis", hdr_font, hdr_fill, alignment=Alignment(horizontal="left", vertical="center"))
ws1.row_dimensions[1].height = 30
for c in range(1, 7):
    ws1.cell(row=1, column=c).fill = hdr_fill
    ws1.cell(row=1, column=c).font = hdr_font

ws1.merge_cells('A2:F2')
set_cell(ws1, 2, 1, "Prepared April 2026 | Baseline: 2 sales/day/store, 384 stores", Font(name="Arial", italic=True, size=9, color="666666"))

r = 4

# ── SECTION A: Key Assumptions ──
set_cell(ws1, r, 1, "A", section_font)
set_cell(ws1, r, 2, "KEY ASSUMPTIONS", section_font)
r += 1

# Headers
for c, h in [(2, "Parameter"), (3, "Value"), (4, "Unit"), (5, "Source")]:
    set_cell(ws1, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
r += 1

# A1: Sales per day per store
set_cell(ws1, r, 2, "Subscriptions sold per store per day", label_font)
set_cell(ws1, r, 3, 2, blue_input, num_fmt=NUM_FMT)  # C6
set_cell(ws1, r, 4, "subs/day", label_font)
set_cell(ws1, r, 5, "Model assumption", Font(name="Arial", italic=True, size=9, color="666666"))
A_SALES_ROW = r
r += 1

# A2: Number of stores
set_cell(ws1, r, 2, "Number of ADNOC Oasis stores", label_font)
set_cell(ws1, r, 3, 384, blue_input, num_fmt=NUM_FMT)  # C7
set_cell(ws1, r, 4, "stores", label_font)
set_cell(ws1, r, 5, "ADNOC counter-proposal", Font(name="Arial", italic=True, size=9, color="666666"))
A_STORES_ROW = r
r += 1

# A3: Tier mix
set_cell(ws1, r, 2, "1-month tier mix", label_font)
set_cell(ws1, r, 3, 0.30, blue_input, num_fmt=PCT_FMT)  # C8
A_MIX1_ROW = r
r += 1
set_cell(ws1, r, 2, "3-month tier mix", label_font)
set_cell(ws1, r, 3, 0.65, blue_input, num_fmt=PCT_FMT)  # C9
A_MIX3_ROW = r
r += 1
set_cell(ws1, r, 2, "12-month tier mix", label_font)
set_cell(ws1, r, 3, 0.05, blue_input, num_fmt=PCT_FMT)  # C10
A_MIX12_ROW = r
r += 1

# A4: Fill frequency (ADNOC's assumption)
set_cell(ws1, r, 2, "Fill frequency (days between fills)", label_font)
set_cell(ws1, r, 3, 3, blue_input, num_fmt=NUM_FMT)  # C11
set_cell(ws1, r, 4, "days", label_font)
set_cell(ws1, r, 5, "ADNOC assumption", Font(name="Arial", italic=True, size=9, color="666666"))
A_FREQ_ROW = r
r += 2

# ── SECTION B: Subscription Sales Volume ──
set_cell(ws1, r, 1, "B", section_font)
set_cell(ws1, r, 2, "SUBSCRIPTION SALES VOLUME (What Our Model Tracks)", section_font)
r += 1

for c, h in [(2, "Metric"), (3, "Value"), (4, "Formula")]:
    set_cell(ws1, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
r += 1

# B1: Network daily sales
set_cell(ws1, r, 2, "Network daily subscription sales", label_font)
ws1.cell(row=r, column=3).value = f'=C{A_SALES_ROW}*C{A_STORES_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
set_cell(ws1, r, 4, f"= C{A_SALES_ROW} × C{A_STORES_ROW}", Font(name="Arial", italic=True, size=9, color="666666"))
B_DAILY_ROW = r
r += 1

# B2: Monthly sales
set_cell(ws1, r, 2, "Monthly subscription sales (×30)", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_DAILY_ROW}*30'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
B_MONTHLY_ROW = r
r += 1

# B3: Yearly sales
set_cell(ws1, r, 2, "Yearly subscription sales (×360)", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_DAILY_ROW}*360'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
set_cell(ws1, r, 4, "30 days × 12 months (matches Excel model)", Font(name="Arial", italic=True, size=9, color="666666"))
B_YEARLY_ROW = r
r += 1

# Tier breakdown
set_cell(ws1, r, 2, "  → 1-month subscriptions sold/year", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_YEARLY_ROW}*C{A_MIX1_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
B_TIER1_ROW = r
r += 1

set_cell(ws1, r, 2, "  → 3-month subscriptions sold/year", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_YEARLY_ROW}*C{A_MIX3_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
B_TIER3_ROW = r
r += 1

set_cell(ws1, r, 2, "  → 12-month subscriptions sold/year", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_YEARLY_ROW}*C{A_MIX12_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
B_TIER12_ROW = r
r += 2

# ── SECTION C: Active Subscriber Pool ──
set_cell(ws1, r, 1, "C", section_font)
set_cell(ws1, r, 2, "ACTIVE SUBSCRIBER POOL AT ANY GIVEN TIME", section_font)
r += 1

ws1.merge_cells(f'B{r}:F{r}')
set_cell(ws1, r, 2, "Not all subscribers are active simultaneously. A 1-month plan bought in January expires by February.", Font(name="Arial", italic=True, size=9, color="666666"))
r += 1

for c, h in [(2, "Tier"), (3, "Sold/Year"), (4, "Active Duration"), (5, "Avg Active Subs")]:
    set_cell(ws1, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
HDR_C_ROW = r
r += 1

# 1-month: active = sold/year ÷ 12 (1 month out of 12 are active)
set_cell(ws1, r, 2, "1-month tier", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_TIER1_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
set_cell(ws1, r, 4, "1 month (÷12)", label_font)
ws1.cell(row=r, column=5).value = f'=C{r}/12'
ws1.cell(row=r, column=5).font = black_formula
ws1.cell(row=r, column=5).number_format = NUM_FMT
C_ACT1_ROW = r
r += 1

# 3-month: active = sold/year ÷ 4
set_cell(ws1, r, 2, "3-month tier", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_TIER3_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
set_cell(ws1, r, 4, "3 months (÷4)", label_font)
ws1.cell(row=r, column=5).value = f'=C{r}/4'
ws1.cell(row=r, column=5).font = black_formula
ws1.cell(row=r, column=5).number_format = NUM_FMT
C_ACT3_ROW = r
r += 1

# 12-month: ramp-up average ≈ sold/year ÷ 2
set_cell(ws1, r, 2, "12-month tier", label_font)
ws1.cell(row=r, column=3).value = f'=C{B_TIER12_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
set_cell(ws1, r, 4, "12 months (avg ÷2)", label_font)
ws1.cell(row=r, column=5).value = f'=C{r}/2'
ws1.cell(row=r, column=5).font = black_formula
ws1.cell(row=r, column=5).number_format = NUM_FMT
C_ACT12_ROW = r
r += 1

# Total active
set_cell(ws1, r, 2, "TOTAL AVERAGE ACTIVE SUBSCRIBERS", result_font)
ws1.cell(row=r, column=5).value = f'=E{C_ACT1_ROW}+E{C_ACT3_ROW}+E{C_ACT12_ROW}'
ws1.cell(row=r, column=5).font = result_font
ws1.cell(row=r, column=5).number_format = NUM_FMT
ws1.cell(row=r, column=5).fill = highlight_fill
C_TOTAL_ROW = r
r += 1

set_cell(ws1, r, 2, "Active subscribers per store", label_font)
ws1.cell(row=r, column=5).value = f'=E{C_TOTAL_ROW}/C{A_STORES_ROW}'
ws1.cell(row=r, column=5).font = bold_font
ws1.cell(row=r, column=5).number_format = DEC_FMT
C_PER_STORE_ROW = r
r += 2

# ── SECTION D: Fill Volume ──
set_cell(ws1, r, 1, "D", section_font)
set_cell(ws1, r, 2, "FILL VOLUME CALCULATION", section_font)
r += 1

for c, h in [(2, "Metric"), (3, "Value"), (4, "Per Store"), (5, "Formula")]:
    set_cell(ws1, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
r += 1

# Daily fills
set_cell(ws1, r, 2, "Daily fills (network)", label_font)
ws1.cell(row=r, column=3).value = f'=E{C_TOTAL_ROW}/C{A_FREQ_ROW}'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
ws1.cell(row=r, column=4).value = f'=C{r}/C{A_STORES_ROW}'
ws1.cell(row=r, column=4).font = black_formula
ws1.cell(row=r, column=4).number_format = DEC_FMT
set_cell(ws1, r, 5, "Active subs ÷ fill frequency", Font(name="Arial", italic=True, size=9, color="666666"))
D_DAILY_ROW = r
r += 1

# Monthly fills
set_cell(ws1, r, 2, "Monthly fills (network)", label_font)
ws1.cell(row=r, column=3).value = f'=C{D_DAILY_ROW}*30'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = NUM_FMT
ws1.cell(row=r, column=4).value = f'=C{r}/C{A_STORES_ROW}'
ws1.cell(row=r, column=4).font = black_formula
ws1.cell(row=r, column=4).number_format = NUM_FMT
D_MONTHLY_ROW = r
r += 1

# Yearly fills
set_cell(ws1, r, 2, "YEARLY FILLS (NETWORK)", result_font)
ws1.cell(row=r, column=3).value = f'=C{D_DAILY_ROW}*365'
ws1.cell(row=r, column=3).font = result_font
ws1.cell(row=r, column=3).number_format = NUM_FMT
ws1.cell(row=r, column=3).fill = highlight_fill
ws1.cell(row=r, column=4).value = f'=C{r}/C{A_STORES_ROW}'
ws1.cell(row=r, column=4).font = result_font
ws1.cell(row=r, column=4).number_format = NUM_FMT
D_YEARLY_ROW = r
r += 2

# ── SECTION E: Comparison to ADNOC's Claim ──
set_cell(ws1, r, 1, "E", section_font)
set_cell(ws1, r, 2, "COMPARISON TO ADNOC'S CLAIM", section_font)
r += 1

for c, h in [(2, "Metric"), (3, "ADNOC Claim"), (4, "Our Analysis"), (5, "Delta")]:
    set_cell(ws1, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
r += 1

set_cell(ws1, r, 2, "Annual fills", label_font)
set_cell(ws1, r, 3, 29000000, blue_input, num_fmt=NUM_FMT)
ws1.cell(row=r, column=4).value = f'=C{D_YEARLY_ROW}'
ws1.cell(row=r, column=4).font = black_formula
ws1.cell(row=r, column=4).number_format = NUM_FMT
ws1.cell(row=r, column=5).value = f'=C{r}-D{r}'
ws1.cell(row=r, column=5).font = Font(name="Arial", bold=True, color="FF0000", size=10)
ws1.cell(row=r, column=5).number_format = NUM_FMT
E_COMP_ROW = r
r += 1

set_cell(ws1, r, 2, "Overstatement factor", label_font)
ws1.cell(row=r, column=5).value = f'=C{E_COMP_ROW}/D{E_COMP_ROW}'
ws1.cell(row=r, column=5).font = Font(name="Arial", bold=True, color="FF0000", size=11)
ws1.cell(row=r, column=5).number_format = '0.0"×"'
r += 1

set_cell(ws1, r, 2, "Fills per store per day (ADNOC)", label_font)
ws1.cell(row=r, column=3).value = f'=C{E_COMP_ROW}/384/365'
ws1.cell(row=r, column=3).font = black_formula
ws1.cell(row=r, column=3).number_format = DEC_FMT
set_cell(ws1, r, 4, "", label_font)
set_cell(ws1, r, 5, "← Implausible: 207 fills/store/day", Font(name="Arial", italic=True, size=9, color="FF0000"))
r += 1

set_cell(ws1, r, 2, "Fills per store per day (Ours)", label_font)
ws1.cell(row=r, column=4).value = f'=D{D_DAILY_ROW}'
ws1.cell(row=r, column=4).font = black_formula
ws1.cell(row=r, column=4).number_format = DEC_FMT
r += 2

# ── SECTION F: Sensitivity — Fill Frequency ──
set_cell(ws1, r, 1, "F", section_font)
set_cell(ws1, r, 2, "SENSITIVITY: FILL FREQUENCY", section_font)
r += 1

for c, h in [(2, "Fill Frequency"), (3, "Daily Fills"), (4, "Annual Fills"), (5, "vs ADNOC 29M")]:
    set_cell(ws1, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
F_HDR_ROW = r
r += 1

freqs = [
    ("Every day", 1),
    ("Every 2 days", 2),
    ("Every 3 days (ADNOC assumption)", 3),
    ("Every 5 days", 5),
    ("Once per week", 7),
]

for label, days in freqs:
    set_cell(ws1, r, 2, label, bold_font if days == 3 else label_font)
    set_cell(ws1, r, 3, None)  # placeholder col for freq value stored nearby
    # Store frequency in col F (hidden helper)
    ws1.cell(row=r, column=6).value = days

    ws1.cell(row=r, column=3).value = f'=E{C_TOTAL_ROW}/F{r}'
    ws1.cell(row=r, column=3).font = black_formula
    ws1.cell(row=r, column=3).number_format = NUM_FMT

    ws1.cell(row=r, column=4).value = f'=C{r}*365'
    ws1.cell(row=r, column=4).font = black_formula
    ws1.cell(row=r, column=4).number_format = NUM_FMT

    ws1.cell(row=r, column=5).value = f'=D{r}/C{E_COMP_ROW}'
    ws1.cell(row=r, column=5).font = black_formula
    ws1.cell(row=r, column=5).number_format = PCT_FMT

    if days == 3:
        for c in range(2, 6):
            ws1.cell(row=r, column=c).fill = highlight_fill

    r += 1

r += 1
ws1.merge_cells(f'B{r}:F{r}')
set_cell(ws1, r, 2, "Even at the most aggressive assumption (filling every single day), annual fills are 21.4M — still below ADNOC's 29M claim.", Font(name="Arial", bold=True, italic=True, size=10, color="FF0000"))
r += 1

# Hide helper column F
ws1.column_dimensions['F'].width = 8

# ════════════════════════════════════════
# SHEET 2: ADNOC Math Check
# ════════════════════════════════════════
ws2 = wb.create_sheet("ADNOC Math Check")
ws2.sheet_properties.tabColor = ADNOC_BLUE

ws2.column_dimensions['A'].width = 4
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 22

# Title
ws2.merge_cells('A1:E1')
set_cell(ws2, 1, 1, "Reverse-Engineering ADNOC's 29M Fills Claim", hdr_font, hdr_fill, alignment=Alignment(horizontal="left", vertical="center"))
ws2.row_dimensions[1].height = 30
for c in range(1, 6):
    ws2.cell(row=1, column=c).fill = hdr_fill

r = 3
set_cell(ws2, r, 1, "A", section_font)
set_cell(ws2, r, 2, "WHAT SUBSCRIBER COUNT PRODUCES 29M FILLS?", section_font)
r += 1

for c, h in [(2, "Working"), (3, "Value"), (4, "Unit")]:
    set_cell(ws2, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
r += 1

set_cell(ws2, r, 2, "ADNOC's claimed fills per year", label_font)
set_cell(ws2, r, 3, 29000000, blue_input, num_fmt=NUM_FMT)
R2_FILLS = r
r += 1

set_cell(ws2, r, 2, "Fill frequency (days between fills)", label_font)
set_cell(ws2, r, 3, 3, blue_input, num_fmt=NUM_FMT)
set_cell(ws2, r, 4, "days", label_font)
R2_FREQ = r
r += 1

set_cell(ws2, r, 2, "Fills per subscriber per year", label_font)
ws2.cell(row=r, column=3).value = f'=ROUND(365/C{R2_FREQ},0)'
ws2.cell(row=r, column=3).font = black_formula
ws2.cell(row=r, column=3).number_format = NUM_FMT
R2_FPY = r
r += 1

set_cell(ws2, r, 2, "Required active subscribers to hit 29M fills", result_font)
ws2.cell(row=r, column=3).value = f'=ROUND(C{R2_FILLS}/C{R2_FPY},0)'
ws2.cell(row=r, column=3).font = result_font
ws2.cell(row=r, column=3).number_format = NUM_FMT
ws2.cell(row=r, column=3).fill = highlight_fill
R2_REQ = r
r += 1

set_cell(ws2, r, 2, "Number of stores", label_font)
set_cell(ws2, r, 3, 384, blue_input, num_fmt=NUM_FMT)
R2_STORES = r
r += 1

set_cell(ws2, r, 2, "Required active subscribers PER STORE", result_font)
ws2.cell(row=r, column=3).value = f'=ROUND(C{R2_REQ}/C{R2_STORES},0)'
ws2.cell(row=r, column=3).font = Font(name="Arial", bold=True, color="FF0000", size=12)
ws2.cell(row=r, column=3).number_format = NUM_FMT
ws2.cell(row=r, column=3).fill = highlight_fill
R2_PER_STORE = r
r += 1

set_cell(ws2, r, 2, "Our model's active subscribers per store", label_font)
ws2.cell(row=r, column=3).value = f'=\'Fills Buildup\'!E{C_PER_STORE_ROW}'
ws2.cell(row=r, column=3).font = Font(name="Arial", color="008000", size=10)
ws2.cell(row=r, column=3).number_format = DEC_FMT
R2_OURS = r
r += 1

ws2.merge_cells(f'B{r}:E{r}')
set_cell(ws2, r, 2, "", Font(name="Arial", bold=True, italic=True, size=10, color="FF0000"))
ws2.cell(row=r, column=2).value = f'=CONCATENATE("ADNOC needs ",TEXT(C{R2_PER_STORE},"#,##0")," active subscribers/store vs our ",TEXT(C{R2_OURS},"#,##0.0")," — a ",TEXT(C{R2_PER_STORE}/C{R2_OURS},"#,##0.0"),"× difference")'
ws2.cell(row=r, column=2).font = Font(name="Arial", bold=True, italic=True, size=10, color="FF0000")
r += 2

# Section B: What does 2.1 subs/store actually mean?
set_cell(ws2, r, 1, "B", section_font)
set_cell(ws2, r, 2, 'WHAT DOES "2.1 SUBSCRIBERS PER STORE" ACTUALLY PRODUCE?', section_font)
r += 1

for c, h in [(2, "Working"), (3, "Value"), (4, "Unit")]:
    set_cell(ws2, r, c, h, hdr_font, hdr_fill, alignment=hdr_align)
r += 1

set_cell(ws2, r, 2, 'ADNOC\'s stated "subscribers per store"', label_font)
set_cell(ws2, r, 3, 2.1, blue_input, num_fmt=DEC_FMT)
R2B_SUBS = r
r += 1

set_cell(ws2, r, 2, "Total subscribers (2.1 × 384)", label_font)
ws2.cell(row=r, column=3).value = f'=C{R2B_SUBS}*C{R2_STORES}'
ws2.cell(row=r, column=3).font = black_formula
ws2.cell(row=r, column=3).number_format = DEC_FMT
R2B_TOTAL = r
r += 1

set_cell(ws2, r, 2, "Fills per year (at once every 3 days)", label_font)
ws2.cell(row=r, column=3).value = f'=C{R2B_TOTAL}*C{R2_FPY}'
ws2.cell(row=r, column=3).font = result_font
ws2.cell(row=r, column=3).number_format = NUM_FMT
ws2.cell(row=r, column=3).fill = highlight_fill
R2B_FILLS = r
r += 1

ws2.merge_cells(f'B{r}:E{r}')
ws2.cell(row=r, column=2).value = f'=CONCATENATE("2.1 subs/store × 384 stores × 122 fills/year = ",TEXT(C{R2B_FILLS},"#,##0")," fills — NOT 29 million")'
ws2.cell(row=r, column=2).font = Font(name="Arial", bold=True, italic=True, size=10, color="FF0000")
r += 2

# Section C: Implied fills per subscriber to get 29M from 2.1/store
set_cell(ws2, r, 1, "C", section_font)
set_cell(ws2, r, 2, "ABSURDITY CHECK: FILLS NEEDED PER SUBSCRIBER", section_font)
r += 1

set_cell(ws2, r, 2, "Total subscribers at 2.1/store", label_font)
ws2.cell(row=r, column=3).value = f'=C{R2B_TOTAL}'
ws2.cell(row=r, column=3).font = black_formula
ws2.cell(row=r, column=3).number_format = DEC_FMT
R2C_TOTAL = r
r += 1

set_cell(ws2, r, 2, "Fills per subscriber per year to reach 29M", label_font)
ws2.cell(row=r, column=3).value = f'=C{R2_FILLS}/C{R2C_TOTAL}'
ws2.cell(row=r, column=3).font = Font(name="Arial", bold=True, color="FF0000", size=12)
ws2.cell(row=r, column=3).number_format = NUM_FMT
R2C_FPY = r
r += 1

set_cell(ws2, r, 2, "That's fills per day per subscriber", label_font)
ws2.cell(row=r, column=3).value = f'=C{R2C_FPY}/365'
ws2.cell(row=r, column=3).font = Font(name="Arial", bold=True, color="FF0000", size=12)
ws2.cell(row=r, column=3).number_format = DEC_FMT
r += 1

ws2.merge_cells(f'B{r}:E{r}')
set_cell(ws2, r, 2, "Each subscriber would need to fill ~99 bottles PER DAY — clearly a math error.", Font(name="Arial", bold=True, italic=True, size=10, color="FF0000"))

# Add borders to all data cells
for ws in [ws1, ws2]:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

# Print settings
for ws in [ws1, ws2]:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

output = "/Users/sami/Desktop/Claude Code/adnoc-model/ADNOC_Fills_Analysis.xlsx"
wb.save(output)
print(f"Saved to {output}")
